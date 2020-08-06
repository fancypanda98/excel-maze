from openpyxl.styles import Side



def solveMaze(path, row, col):
    while len(path) > 0:
        if path[-1].getRow() == row and path[-1].getColumn() == col:
            return path
        while len(path) > 0 and len(path[-1].getNodes()) == 0:
            path.pop()
        node = path[-1].getLast()
        path[-1].popNode()
        path.append(node)


def determineValidCells(page, node, prev):
    ret = []
    S = Side(None, None, None)

    if page.cell(node.getRow(), node.getColumn()).border.bottom == S:
        if not (node.getColumn() == prev.getColumn() and node.getRow() + 1 == prev.getRow()):
            app = MazeTreeNode(node.getRow() + 1, node.getColumn())
            ret.append(app)
    if page.cell(node.getRow(), node.getColumn()).border.top == S:
        if not (node.getColumn() == prev.getColumn() and node.getRow() - 1 == prev.getRow()):
            app = MazeTreeNode(node.getRow() - 1, node.getColumn())
            ret.append(app)
    if page.cell(node.getRow(), node.getColumn()).border.right == S:
        if not (node.getColumn() + 1 == prev.getColumn() and node.getRow() == prev.getRow()):
            app = MazeTreeNode(node.getRow(), node.getColumn() + 1)
            ret.append(app)
    if page.cell(node.getRow(), node.getColumn()).border.left == S:
        if not (node.getColumn() - 1 == prev.getColumn() and node.getRow() == prev.getRow()):
            app = MazeTreeNode(node.getRow(), node.getColumn() - 1)
            ret.append(app)
    return ret


# keep parent list and check with each parent if the node being processed is the last node it has, then pop,
# dont put dead end nodes on the parent list
def generateTreeFromMaze(page, node, prev, row, col):
    queue = [node]
    parent = [prev]
    while len(queue) > 0:
        n = determineValidCells(page, queue[0], parent[0])
        for i in n:
            queue[0].insertNode(i)
            queue.append(i)
        if len(queue[0].getNodes()) > 0:
            parent.append(queue[0])
        if len(parent) > 0:
            if parent[0].getLast().getColumn() == queue[0].getColumn() and parent[0].getLast().getRow() == queue[0].getRow():
                temp = parent.pop(0)
                if temp.getRow() == row and temp.getColumn() == col:
                    return node
        queue.pop(0)
    return node


class MazeTreeNode(object):

    def __init__(self, row=-1, column=-1):
        self.nodes = []
        self.column = column
        self.row = row

    def getNodes(self):
        return self.nodes

    def getRow(self):
        return self.row

    def getColumn(self):
        return self.column

    def insertNode(self, node):
        self.nodes.append(node)

    def setCell(self, row, column):
        self.row = row
        self.column = column

    def popNode(self):
        self.nodes.pop()

    def getLast(self):
        return self.nodes[-1]
