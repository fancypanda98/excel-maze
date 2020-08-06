import openpyxl
from openpyxl.styles import PatternFill

from tree import MazeTreeNode
from tree import solveMaze
from tree import generateTreeFromMaze



wb = openpyxl.load_workbook("Maze.xlsx")
sheet = wb.active

height = sheet['A1'].value
width = sheet['E1'].value
endingCell = sheet.cell(height+6, 4)
endingCol = sheet[endingCell.value].column
endingRow = sheet[endingCell.value].row

if width <= 0 or height <= 0:
    print("Maze Dimensions too Small")
    exit()

startingCell = MazeTreeNode(sheet[sheet.cell(height+8,4).value].row, sheet[sheet.cell(height+8,4).value].column)
nullCell = MazeTreeNode()
nullCell.insertNode(startingCell)

tree = generateTreeFromMaze(sheet, startingCell, nullCell, endingRow, endingCol)

print("Maze Analyzed")

solution = [tree]
solution = solveMaze(solution, endingRow, endingCol)

print("Maze Solved")

whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
greenFill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

for i in range(4, height+4):
    for j in range(4, width+4):
        sheet.cell(i, j).fill = whiteFill

for i in solution:
    sheet.cell(i.getRow(), i.getColumn()).fill = greenFill

print("Path Drawn")

wb.save("Maze.xlsx")

print("Maze Saved")
print()
print()
input("Press Enter to Continue")
