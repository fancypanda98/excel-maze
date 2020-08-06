from random import randint
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

from disjointset import DisjointSet


class box:
    left = True
    right = True
    up = True
    down = True


height =int(input("Enter Maze Height:   "))
width = int(input("Enter Maze Width :   "))
print()
current = 1

board = []
walls = []
handler = DisjointSet(height * width)
# True means there is a wall there for walls
# True means we can tear down the wall for board
for i in range(0, height * width):
    temp = box()
    temp.left = False
    temp.right = False
    temp.down = False
    temp.up = False
    temp2 = box()
    board.append(temp)
    walls.append(temp2)
    print("Preparing Maze            :  ", "{:.{}f}".format((((i+1) / (width * height)) * 100), 4), "%", end="\r")

    if not (0 <= i <= width):
        board[i].up = True
    if not (i % width == 0):
        board[i].left = True
    if not ((i % width) == (width - 1)):
        board[i].right = True
    if not (((height - 1) * width) <= i < (height * width)):
        board[i].down = True

print()

workingCells = []

for i in range(0, height * width):
    workingCells.append(i)

#Creating Random Maze in Disjoint Set and Workbook
while handler.get_num_sets() > 1:
    r = randint(0, len(workingCells)-1)
    index = workingCells[r]
    direction = randint(0, 3)
    count = 0
    stillStanding = True
    while stillStanding:
        if direction == 0 and board[index].up and (not handler.are_in_same_set(index, index - width)):
            stillStanding = False
            otherIndex = index - width
            board[index].up = False
            board[otherIndex].down = False
            walls[index].up = False
            walls[otherIndex].down = False
            handler.merge_sets(index, otherIndex)
        elif direction == 1 and board[index].down and (not handler.are_in_same_set(index, index + width)):
            stillStanding = False
            otherIndex = index + width
            board[index].down = False
            board[otherIndex].up = False
            walls[index].down = False
            walls[otherIndex].up = False
            handler.merge_sets(index, otherIndex)
        elif direction == 2 and board[index].left and (not handler.are_in_same_set(index, index - 1)):
            stillStanding = False
            otherIndex = index - 1
            board[index].left = False
            board[otherIndex].right = False
            walls[index].left = False
            walls[otherIndex].right = False
            handler.merge_sets(index, otherIndex)
        elif direction == 3 and board[index].right and (not handler.are_in_same_set(index, index + 1)):
            stillStanding = False
            otherIndex = index + 1
            board[index].right = False
            board[otherIndex].left = False
            walls[index].right = False
            walls[otherIndex].left = False
            handler.merge_sets(index, otherIndex)
        else:
            if count == 3:
                count = 0
                workingCells.pop(r)
                if r == len(workingCells):
                    r = 0
                index = workingCells[r]
            else:
                count += 1
                direction = (direction + 1) % 4
    current += 1
    print("Creating Maze             :  ", "{:.{}f}".format(((current / ((width * height)))*100), 4), "%", end="\r")

print()

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "By Jacque Keener"
whiteFill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

for i in range(0, height * width):
    L = Side(None, None, None)
    D = Side(None, None, None)
    U = Side(None, None, None)
    R = Side(None, None, None)

    if walls[i].left:
        L = Side(None, None, 'medium')
    if walls[i].right:
        R = Side(None, None, 'medium')
    if walls[i].down:
        D = Side(None, None, 'medium')
    if walls[i].up:
        U = Side(None, None, 'medium')

    sheet.cell((i // width) + 4, (i % width) + 4).border = Border(L, R, U, D)
    sheet.cell((i // width) + 4, (i % width) + 4).fill = whiteFill

    print("Drawing Maze              :  ", "{:.{}f}".format((((i+1) / (width * height)) * 100), 4), "%", end="\r")

sheet.cell(2, 1).value = 'Height'
sheet.cell(1, 1).value = height

sheet.cell(2, 5).value = 'Width'
sheet.cell(1, 5).value = width

sheet.cell((height+5), 4).value = 'Enter Ending Cell Below'
sheet.cell((height+6), 4).border = Border(Side(None, None, None), Side(None, None, None), Side(None, None, None), Side(None, None, 'thick'))
sheet.cell((height+6), 4).value = sheet.cell(height+3, width+3).coordinate

sheet.cell((height+7), 4).value = 'Enter Starting Cell Below'
sheet.cell((height+8), 4).border = Border(Side(None, None, None), Side(None, None, None), Side(None, None, None), Side(None, None, 'thick'))
sheet.cell((height+8), 4).value = sheet.cell(4, 4).coordinate

print()

for i in range(0, width+8):
    c = sheet.cell(i+1, i+1)
    letter = get_column_letter(i+1)
    sheet.column_dimensions[letter].width = 4.0
    print("Adjusting Column Size     :  ", "{:.{}f}".format((((i + 1) / (width+8)) * 100), 4), "%", end="\r")

print()


for i in range(0, height+3):
    c = sheet.cell(i+1, i+1)
    sheet.row_dimensions[i].height = 15.0
    print("Adjusting Row Size        :  ", "{:.{}f}".format((((i + 1) / (height+3)) * 100), 4), "%", end="\r")

print()
print("Saving Maze to Excel")

wb.save("Maze.xlsx")

print("Maze Complete")
print()
print()
input("Press Enter to Continue")
